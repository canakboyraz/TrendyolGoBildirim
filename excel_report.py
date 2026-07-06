"""
Excel rapor motoru — günlük, haftalık, aylık.
Saatlik yoğunluk grafiği ve önceki gün karşılaştırması içerir.
"""
import os, json, base64, requests
from datetime import datetime, date, time as dtime, timedelta
from collections import defaultdict
import pytz, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from dotenv import load_dotenv
load_dotenv()

from config import (SUPPLIER_ID, API_KEY, API_SECRET, INTEGRATOR_NAME,
                    API_BASE_URL, TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID)

TURKEY_TZ = pytz.timezone("Europe/Istanbul")

# ── Renkler ──────────────────────────────────────────────────────────────────
CLR_HEADER     = "1F3864"; CLR_HEADER_F = "FFFFFF"
CLR_SUBHDR     = "2E75B6"; CLR_SUBHDR_F = "FFFFFF"
CLR_ODD        = "EBF3FB"; CLR_EVEN     = "FFFFFF"
CLR_CANCEL     = "FFD7D7"; CLR_TOTAL    = "FFF2CC"
CLR_SUM_HDR    = "375623"; CLR_SUM_F    = "FFFFFF"
CLR_GREEN_HDR  = "1E7145"; CLR_GREEN_F  = "FFFFFF"
CLR_ORANGE_HDR = "C55A11"; CLR_ORANGE_F = "FFFFFF"

# ── Stil yardımcıları ─────────────────────────────────────────────────────────
def _b(style="thin"): s=Side(style=style); return Border(left=s,right=s,top=s,bottom=s)
def _f(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10): return Font(bold=bold,color=color,size=size,name="Calibri")
def _al(h="center",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def _cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w


def _hdr(ws, row, col, val, bg, fg, size=10, span=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = _font(True, fg, size); cell.fill = _f(bg)
    cell.alignment = _al(); cell.border = _b()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    return cell


def _cell(ws, row, col, val, bg=CLR_EVEN, fmt=None, h="center", wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = _f(bg); c.border = _b()
    c.font = _font(size=9); c.alignment = _al(h=h, wrap=wrap)
    if fmt: c.number_format = fmt
    return c

# ── API yardımcıları ──────────────────────────────────────────────────────────
def _headers():
    cred = f"{API_KEY}:{API_SECRET}"
    enc  = base64.b64encode(cred.encode()).decode()
    return {"Authorization": f"Basic {enc}",
            "User-Agent": f"{SUPPLIER_ID} - {INTEGRATOR_NAME}",
            "x-agentname": INTEGRATOR_NAME,
            "x-executor-user": "integration@selfservice.com",
            "Content-Type": "application/json"}


def _range_ms(d: date):
    start = TURKEY_TZ.localize(datetime.combine(d, dtime.min))
    end   = TURKEY_TZ.localize(datetime.combine(d, dtime.max))
    return int(start.timestamp()*1000), int(end.timestamp()*1000)


def fetch_orders(start_ms: int, end_ms: int) -> list:
    all_orders, page = [], 0
    statuses = "Created,Picking,Invoiced,Shipped,Delivered,Cancelled,UnSupplied"
    while True:
        url = f"{API_BASE_URL}/integrator/order/meal/suppliers/{SUPPLIER_ID}/packages"
        params = {"packageStatuses": statuses,
                  "packageModificationStartDate": start_ms,
                  "packageModificationEndDate": end_ms,
                  "page": page, "size": 50}
        try:
            r = requests.get(url, headers=_headers(), params=params, timeout=15)
            r.raise_for_status(); data = r.json()
        except Exception as e:
            print(f"[EXCEL] API hata: {e}"); break
        all_orders.extend(data.get("content", []))
        if page >= data.get("totalPages", 1) - 1: break
        page += 1
    return all_orders

# ── Dönüştürücüler ────────────────────────────────────────────────────────────
CANCELLED_ST = {"Cancelled", "UnSupplied"}

def _status(s): return {"Created":"Yeni","Picking":"Hazırlanıyor","Invoiced":"Hazır",
    "Shipped":"Yolda","Delivered":"Teslim","Cancelled":"İptal","UnSupplied":"Rest. İptali"}.get(s,s)

def _payment(order):
    p = order.get("payment",{}) or {}
    raw = p.get("paymentType","")
    m = {"PAY_WITH_CARD":"Online Kart","PAY_WITH_ON_DELIVERY":"Kapıda Ödeme","PAY_WITH_MEAL_CARD":"Yemek Kartı"}
    label = m.get(raw, raw)
    if raw == "PAY_WITH_ON_DELIVERY":
        sub = (p.get("onDelivery") or {}).get("paymentType","")
        sub_labels = {"CASH": "Nakit", "CARD": "Kart"}
        label += f" - {sub_labels.get(sub, sub)}" if sub else ""
    elif raw == "PAY_WITH_MEAL_CARD":
        ct = (p.get("mealCard") or {}).get("cardSourceType","")
        label += f" ({ct})" if ct else ""
    return label

def _app(a): return {"Trendyol":"Trendyol","TrendyolGo":"TGo","Galaxy":"Getir Yemek"}.get(a or "",a or "-")
def _delivery(d): return {"GO":"TGo Kuryesi","STORE":"Rest. Kuryesi"}.get(d,d)

def _ts_str(ts):
    if not ts: return "-"
    try: return datetime.fromtimestamp(ts/1000, tz=TURKEY_TZ).strftime("%H:%M")
    except: return "-"

def _products_str(order):
    parts = []
    for ln in order.get("lines",[]):
        qty  = len(ln.get("items",[])) or 1
        mods = [m.get("name","") for m in ln.get("modifierProducts",[])]
        mod  = f" ({', '.join(mods)})" if mods else ""
        parts.append(f"{ln.get('name','?')}{mod} x{qty}")
    return "\n".join(parts)

def _order_qty(order):
    return sum(len(ln.get("items",[])) or 1 for ln in order.get("lines",[]))

def _line_revenue(order):
    return sum(ln.get("price",0)*(len(ln.get("items",[])) or 1) for ln in order.get("lines",[]))

# ── Sayfa 1: Sipariş Listesi ──────────────────────────────────────────────────
def _sheet_orders(wb, orders: list, title: str):
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:N1")
    c = ws["A1"]; c.value = title
    c.font=_font(True,CLR_HEADER_F,13); c.fill=_f(CLR_HEADER); c.alignment=_al(); ws.row_dimensions[1].height=28

    cols = [("Sipariş No",16),("Kod",8),("Saat",10),("Statü",16),("Ürünler",42),
            ("Adet",6),("Ürün (₺)",14),("Toplam (₺)",13),("Ödeme",22),
            ("Teslimat",16),("Kaynak",13),("Müşteri",15),("Not",26),("Test",7)]
    for ci,(cn,cw) in enumerate(cols,1):
        _hdr(ws,2,ci,cn,CLR_SUBHDR,CLR_SUBHDR_F); _cw(ws,ci,cw)
    ws.row_dimensions[2].height=18; ws.freeze_panes="A3"

    row=3
    for o in orders:
        st = o.get("packageStatus","")
        bg = CLR_CANCEL if st in CANCELLED_ST else (CLR_ODD if row%2==1 else CLR_EVEN)
        vals = [o.get("orderNumber","-"), o.get("orderCode","-"),
                _ts_str(o.get("packageCreationDate")), _status(st),
                _products_str(o), _order_qty(o), round(_line_revenue(o),2),
                round(o.get("totalPrice",0),2), _payment(o), _delivery(o.get("deliveryType","")),
                _app(o.get("userInformation",{}).get("appName")),
                f"{o.get('customer',{}).get('firstName','')} {o.get('customer',{}).get('lastName','')}".strip(),
                o.get("customerNote","") or "", "Evet" if o.get("testPackage") else "Hayır"]
        for ci,v in enumerate(vals,1):
            _cell(ws,row,ci,v,bg,
                  fmt=('#,##0.00 ₺' if ci in(7,8) else None),
                  h=("left" if ci in(5,9,10,11,13) else "center"),
                  wrap=(ci in(5,13)))
        ws.row_dimensions[row].height = max(15, len(o.get("lines",[])) * 14)
        row+=1

    # Toplam satırı
    active = [o for o in orders if o.get("packageStatus") not in CANCELLED_ST]
    total_rev = sum(o.get("totalPrice",0) for o in active)
    ws.merge_cells(f"A{row}:G{row}")
    c2 = ws.cell(row=row,column=1,value="TOPLAM (İptal hariç)")
    c2.font=_font(True,size=10); c2.fill=_f(CLR_TOTAL); c2.alignment=_al("right"); c2.border=_b()
    tv = ws.cell(row=row,column=8,value=round(total_rev,2))
    tv.font=_font(True,size=10); tv.fill=_f(CLR_TOTAL); tv.alignment=_al(); tv.border=_b(); tv.number_format='#,##0.00 ₺'
    for ci in range(9,15):
        ws.cell(row=row,column=ci).fill=_f(CLR_TOTAL); ws.cell(row=row,column=ci).border=_b()
    return ws

# ── Sayfa 2: Ürün Özeti ───────────────────────────────────────────────────────
def _sheet_products(wb, orders: list, title="Ürün Özeti"):
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:E1")
    c=ws["A1"]; c.value=title
    c.font=_font(True,CLR_HEADER_F,12); c.fill=_f(CLR_HEADER); c.alignment=_al(); ws.row_dimensions[1].height=24

    for ci,(cn,cw) in enumerate([("Ürün Adı",42),("Adet",10),("Toplam (₺)",18),("Ort. Fiyat (₺)",18),("% Pay",10)],1):
        _hdr(ws,2,ci,cn,CLR_SUBHDR,CLR_SUBHDR_F); _cw(ws,ci,cw)

    active = [o for o in orders if o.get("packageStatus") not in CANCELLED_ST]
    pq, pr = defaultdict(int), defaultdict(float)
    for o in active:
        for ln in o.get("lines",[]):
            n=ln.get("name","?"); q=len(ln.get("items",[])) or 1
            pq[n]+=q; pr[n]+=ln.get("price",0)*q

    grand = sum(pr.values()) or 1
    for ri,(name,qty) in enumerate(sorted(pq.items(),key=lambda x:x[1],reverse=True),3):
        rev=pr[name]; avg=rev/qty if qty else 0; pct=rev/grand*100
        bg=CLR_ODD if ri%2==0 else CLR_EVEN
        _cell(ws,ri,1,name,bg,h="left")
        _cell(ws,ri,2,qty,bg)
        _cell(ws,ri,3,round(rev,2),bg,'#,##0.00 ₺')
        _cell(ws,ri,4,round(avg,2),bg,'#,##0.00 ₺')
        _cell(ws,ri,5,round(pct,1),bg,'0.0"%"')

    tr=len(pq)+3
    for ci,val in enumerate([("TOPLAM",""),sum(pq.values()),round(grand,2),"",""],1):
        c2=ws.cell(row=tr,column=ci,value=val[1] if isinstance(val,tuple) else val)
        c2.fill=_f(CLR_TOTAL); c2.border=_b(); c2.font=_font(True,size=10); c2.alignment=_al()
        if ci==3: c2.number_format='#,##0.00 ₺'
    ws.cell(row=tr,column=1).value="TOPLAM"; ws.cell(row=tr,column=1).alignment=_al()
    return ws

# ── Sayfa 3: Saatlik Yoğunluk + Grafik ───────────────────────────────────────
def _sheet_hourly(wb, orders: list, title="Saatlik Yoğunluk"):
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:D1")
    c=ws["A1"]; c.value=title
    c.font=_font(True,CLR_HEADER_F,12); c.fill=_f(CLR_GREEN_HDR); c.alignment=_al(); ws.row_dimensions[1].height=24

    for ci,(cn,cw) in enumerate([("Saat",12),("Sipariş Adedi",16),("Toplam Ciro (₺)",18),("Ort. Sipariş (₺)",18)],1):
        _hdr(ws,2,ci,cn,CLR_GREEN_HDR,CLR_GREEN_F); _cw(ws,ci,cw)

    active = [o for o in orders if o.get("packageStatus") not in CANCELLED_ST]
    h_count = defaultdict(int); h_rev = defaultdict(float)
    for o in active:
        ts = o.get("packageCreationDate",0)
        if ts:
            hour = datetime.fromtimestamp(ts/1000, tz=TURKEY_TZ).hour
            h_count[hour]+=1; h_rev[hour]+=o.get("totalPrice",0)

    data_start = 3
    for ri,hour in enumerate(range(24), data_start):
        cnt = h_count.get(hour,0)
        rev = h_rev.get(hour,0.0)
        avg = rev/cnt if cnt else 0
        bg  = CLR_ODD if ri%2==0 else CLR_EVEN
        _cell(ws,ri,1,f"{hour:02d}:00 - {hour:02d}:59",bg,h="left")
        _cell(ws,ri,2,cnt,bg)
        _cell(ws,ri,3,round(rev,2),bg,'#,##0.00 ₺')
        _cell(ws,ri,4,round(avg,2),bg,'#,##0.00 ₺')

    # Grafik
    chart = BarChart()
    chart.type = "col"; chart.title = title
    chart.y_axis.title = "Sipariş Adedi"; chart.x_axis.title = "Saat"
    chart.style = 10; chart.width = 22; chart.height = 12

    data_ref  = Reference(ws, min_col=2, min_row=data_start, max_row=data_start+23)
    cats_ref  = Reference(ws, min_col=1, min_row=data_start, max_row=data_start+23)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "F3")
    return ws

# ── Sayfa 4: Önceki Gün Karşılaştırması ─────────────────────────────────────
def _sheet_compare(wb, today_orders: list, yesterday_orders: list,
                   today_label: str, yesterday_label: str, title="Karşılaştırma"):
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:D1")
    c=ws["A1"]; c.value=f"{yesterday_label} vs {today_label}"
    c.font=_font(True,CLR_HEADER_F,12); c.fill=_f(CLR_ORANGE_HDR); c.alignment=_al(); ws.row_dimensions[1].height=24

    for ci,(cn,cw) in enumerate([("Metrik",28),(yesterday_label,18),(today_label,18),("Değişim",14)],1):
        _hdr(ws,2,ci,cn,CLR_ORANGE_HDR,CLR_ORANGE_F); _cw(ws,ci,cw)

    def stats(orders):
        active = [o for o in orders if o.get("packageStatus") not in CANCELLED_ST]
        cancelled = [o for o in orders if o.get("packageStatus") in CANCELLED_ST]
        rev  = sum(o.get("totalPrice",0) for o in active)
        qty  = sum(_order_qty(o) for o in active)
        avg  = rev/len(active) if active else 0
        return {"Sipariş Adedi":len(active), "İptal Adedi":len(cancelled),
                "Toplam Ciro (₺)":round(rev,2), "Ort. Sipariş (₺)":round(avg,2),
                "Toplam Ürün Adedi":qty}

    st_y = stats(yesterday_orders)
    st_t = stats(today_orders)
    fmt_map = {"Toplam Ciro (₺)":'#,##0.00 ₺',"Ort. Sipariş (₺)":'#,##0.00 ₺'}

    for ri,(key,vy) in enumerate(st_y.items(), 3):
        vt  = st_t.get(key,0)
        diff = vt - vy
        pct  = (diff/vy*100) if vy else 0
        diff_str = f"+{pct:.1f}%" if pct>0 else f"{pct:.1f}%"
        bg = CLR_ODD if ri%2==0 else CLR_EVEN
        fmt = fmt_map.get(key)
        _cell(ws,ri,1,key,bg,h="left")
        _cell(ws,ri,2,vy,bg,fmt)
        _cell(ws,ri,3,vt,bg,fmt)
        dc = ws.cell(row=ri,column=4,value=diff_str)
        dc.fill=_f(bg); dc.border=_b(); dc.font=_font(bold=True,color=("375623" if diff>=0 else "C00000"),size=10)
        dc.alignment=_al()

    # Saatlik karşılaştırma
    row=len(st_y)+4
    ws.merge_cells(f"A{row}:D{row}")
    hc=ws.cell(row=row,column=1,value="Saatlik Sipariş Karşılaştırması")
    hc.font=_font(True,CLR_ORANGE_F,11); hc.fill=_f(CLR_ORANGE_HDR); hc.alignment=_al(); ws.row_dimensions[row].height=20
    row+=1

    for ci,(cn,cw) in enumerate([("Saat",14),(yesterday_label,18),(today_label,18),("Fark",10)],1):
        _hdr(ws,row,ci,cn,CLR_ORANGE_HDR,CLR_ORANGE_F); _cw(ws,ci,cw)
    row+=1

    def hour_counts(orders):
        hc2 = defaultdict(int)
        for o in orders:
            if o.get("packageStatus") in CANCELLED_ST: continue
            ts = o.get("packageCreationDate",0)
            if ts:
                hc2[datetime.fromtimestamp(ts/1000,tz=TURKEY_TZ).hour]+=1
        return hc2

    hcy=hour_counts(yesterday_orders); hct=hour_counts(today_orders)
    for hour in range(24):
        vy2=hcy.get(hour,0); vt2=hct.get(hour,0); diff2=vt2-vy2
        bg=CLR_ODD if row%2==0 else CLR_EVEN
        _cell(ws,row,1,f"{hour:02d}:00",bg,h="left")
        _cell(ws,row,2,vy2,bg)
        _cell(ws,row,3,vt2,bg)
        dc2=ws.cell(row=row,column=4,value=(f"+{diff2}" if diff2>0 else str(diff2)))
        dc2.fill=_f(bg); dc2.border=_b()
        dc2.font=_font(bold=True,color=("375623" if diff2>=0 else "C00000"),size=9)
        dc2.alignment=_al(); row+=1
    return ws

# ── Sayfa 5: Genel Özet ───────────────────────────────────────────────────────
def _sheet_summary(wb, orders: list, title: str, date_label: str):
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=22
    ws.merge_cells("A1:B1")
    c=ws["A1"]; c.value=f"Genel Özet — {date_label}"
    c.font=_font(True,CLR_HEADER_F,12); c.fill=_f(CLR_HEADER); c.alignment=_al(); ws.row_dimensions[1].height=24

    active    = [o for o in orders if o.get("packageStatus") not in CANCELLED_ST]
    cancelled = [o for o in orders if o.get("packageStatus") in CANCELLED_ST]
    revenue   = sum(o.get("totalPrice",0) for o in active)
    avg_order = revenue/len(active) if active else 0

    pay_cnt=defaultdict(int); app_cnt=defaultdict(int)
    for o in active:
        pay_cnt[_payment(o)]+=1
        app_cnt[_app(o.get("userInformation",{}).get("appName",""))]+=1

    rows=[("📦 Toplam Sipariş",len(orders)),("✅ Geçerli",len(active)),
          ("❌ İptal",len(cancelled)),("💰 Toplam Ciro (₺)",round(revenue,2)),
          ("💳 Ort. Sipariş (₺)",round(avg_order,2)),("",""),("— Ödeme —","")]
    for lb,cnt in sorted(pay_cnt.items(),key=lambda x:x[1],reverse=True):
        rows.append((f"  {lb}",cnt))
    rows+=([("",""),("— Kaynak —","")])
    for lb,cnt in sorted(app_cnt.items(),key=lambda x:x[1],reverse=True):
        rows.append((f"  {lb}",cnt))

    for ri,(label,value) in enumerate(rows,2):
        is_sec = str(label).startswith("—")
        is_main= ri in (2,3,4,5,6)
        la=ws.cell(row=ri,column=1,value=label)
        va=ws.cell(row=ri,column=2,value=value)
        if is_sec:
            for c2 in(la,va): c2.fill=_f(CLR_SUM_HDR); c2.font=_font(True,CLR_SUM_F,10)
        elif is_main:
            for c2 in(la,va): c2.fill=_f(CLR_TOTAL); c2.font=_font(True,size=10)
        else:
            bg=CLR_ODD if ri%2==0 else CLR_EVEN
            for c2 in(la,va): c2.fill=_f(bg); c2.font=_font(size=10)
        for c2 in(la,va): c2.border=_b(); c2.alignment=_al(h="left")
        va.alignment=_al()
        if label in("💰 Toplam Ciro (₺)","💳 Ort. Sipariş (₺)"): va.number_format='#,##0.00 ₺'
    return ws

# ── Telegram gönderici ────────────────────────────────────────────────────────
def _send_file(filepath: str, caption: str):
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        with open(filepath,"rb") as f:
            r = requests.post(url,
                data={"chat_id":TELEGRAM_CHAT_ID,"caption":caption,"parse_mode":"HTML"},
                files={"document":(os.path.basename(filepath),f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30)
        r.raise_for_status(); print(f"[EXCEL] ✅ Gönderildi: {os.path.basename(filepath)}")
        return True
    except Exception as e:
        print(f"[EXCEL] ❌ Gönderilemedi: {e}"); return False


# ── Ana rapor fonksiyonları ───────────────────────────────────────────────────
def build_and_send(orders_today, orders_yesterday, filename: str,
                   date_label: str, period_label: str, weekly_data: list = None):
    """Excel oluşturur ve Telegram'a gönderir."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # boş default sayfayı kaldır

    # Haftalık/aylık modda tüm günlerin listesi
    if weekly_data:
        all_orders = [o for day_orders in weekly_data for o in day_orders]
        _sheet_orders(wb, all_orders, f"Siparişler ({period_label})")
        _sheet_products(wb, all_orders, "Ürün Özeti")
        _sheet_hourly(wb, all_orders, "Saatlik Yoğunluk")
        _sheet_summary(wb, all_orders, "Genel Özet", date_label)
    else:
        _sheet_orders(wb, orders_today, f"Siparişler ({date_label})")
        _sheet_products(wb, orders_today, "Ürün Özeti")
        _sheet_hourly(wb, orders_today, "Saatlik Yoğunluk")
        if orders_yesterday is not None:
            yesterday_label = (datetime.now(TURKEY_TZ).date()-timedelta(days=1)).strftime("%d.%m")
            _sheet_compare(wb, orders_today, orders_yesterday,
                           date_label, yesterday_label)
        _sheet_summary(wb, orders_today, "Genel Özet", date_label)

    wb.save(filename)
    caption = f"📊 <b>{period_label} Raporu — {date_label}</b>"
    _send_file(filename, caption)
    if os.path.exists(filename): os.remove(filename)


def generate_daily():
    today = datetime.now(TURKEY_TZ).date()
    yesterday = today - timedelta(days=1)
    label = today.strftime("%d.%m.%Y")
    print(f"[EXCEL] Günlük rapor hazırlanıyor ({label})...")
    s,e = _range_ms(today)
    sy,ey = _range_ms(yesterday)
    orders_today     = fetch_orders(s, e)
    orders_yesterday = fetch_orders(sy, ey)
    print(f"[EXCEL] Bugün: {len(orders_today)}, Dün: {len(orders_yesterday)}")
    fname = f"TGo_Gunluk_{label.replace('.','')}.xlsx"
    build_and_send(orders_today, orders_yesterday, fname, label, "Günlük")


def generate_weekly():
    today = datetime.now(TURKEY_TZ).date()
    # Geçen haftanın Pazartesi → Pazar
    end_day   = today - timedelta(days=today.weekday()+1)   # geçen Pazar
    start_day = end_day - timedelta(days=6)                 # geçen Pazartesi
    label = f"{start_day.strftime('%d.%m')} - {end_day.strftime('%d.%m.%Y')}"
    print(f"[EXCEL] Haftalık rapor hazırlanıyor ({label})...")
    weekly_data = []
    for i in range(7):
        d = start_day + timedelta(days=i)
        s,e = _range_ms(d)
        weekly_data.append(fetch_orders(s,e))
    fname = f"TGo_Haftalik_{start_day.strftime('%d%m')}.xlsx"
    build_and_send(None, None, fname, label, "Haftalık", weekly_data=weekly_data)


def generate_monthly():
    today = datetime.now(TURKEY_TZ).date()
    # Geçen ay
    first_this = today.replace(day=1)
    last_prev  = first_this - timedelta(days=1)
    first_prev = last_prev.replace(day=1)
    label = first_prev.strftime("%B %Y")
    print(f"[EXCEL] Aylık rapor hazırlanıyor ({label})...")
    monthly_data = []
    d = first_prev
    while d <= last_prev:
        s,e = _range_ms(d)
        monthly_data.append(fetch_orders(s,e))
        d += timedelta(days=1)
    fname = f"TGo_Aylik_{first_prev.strftime('%Y%m')}.xlsx"
    build_and_send(None, None, fname, label, "Aylık", weekly_data=monthly_data)


if __name__ == "__main__":
    generate_daily()
